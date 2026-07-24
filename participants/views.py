from django.shortcuts import render, redirect, get_object_or_404
from .models import Participant
from .forms import ParticipantForm
from django.db.models import Q
from openpyxl import load_workbook
from django.http import JsonResponse
# from django.db.models import Count



# def participant_list(request):

#     participants = Participant.objects.all()

#     return render(
#         request,
#         'participants/list.html',
#         {
#             'participants': participants
#         }
#     )


# def dashboard(request):

#     participants = Participant.objects.all()

#     context = {
#         "total_participants": participants.count(),
#         "total_pasteurs": participants.filter(quality="pasteur").count(),
#         "total_responsables": participants.filter(quality="responsable").count(),
#         "total_etudiants": participants.filter(quality="etudiant").count(),

#         "last_participants":
#             participants.order_by("-created_at")[:8],

#         "countries":
#             participants.values("country")
#                         .annotate(total=Count("id"))
#                         .order_by("-total"),
#     }

#     return render(
#         request,
#         "dashboard.html",
#         context
#     )


# @login_required
def participant_import(request):

    if request.method != "POST":

        return JsonResponse({

            "message":"Méthode invalide."

        },status=400)


    fichier = request.FILES.get("excel")

    if not fichier:

        return JsonResponse({

            "message":"Aucun fichier sélectionné."

        },status=400)


    wb = load_workbook(fichier)

    ws = wb.active


    compteur = 0


    for row in ws.iter_rows(min_row=2, values_only=True):

        Participant.objects.create(

            full_name=row[0],

            local_church=row[1],

            congregation=row[2],

            city=row[3],

            country=row[4],

            quality=row[5],

            phone=row[6],

            created_by=request.user

        )

        compteur += 1

    return JsonResponse({

        "message":f"{compteur} participant(s) importé(s)."

    })




def participant_list(request):

    query = request.GET.get('q', '')

    participants = Participant.objects.all().order_by('-created_at')


    if query:

        participants = participants.filter(
            Q(full_name__icontains=query) |
            Q(local_church__icontains=query) |
            Q(congregation__icontains=query) |
            Q(city__icontains=query) |
            Q(country__icontains=query) |
            Q(phone__icontains=query)
        )


    context = {
        'participants': participants,
        'query': query,
    }


    return render(
        request,
        'participants/list.html',
        context
    )



def participant_create(request):

    if request.method == "POST":

        form = ParticipantForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():
            form.save()
            return redirect('participant_list')

    else:
        form = ParticipantForm()


    return render(
        request,
        'participants/form.html',
        {
            'form': form
        }
    )





def participant_delete(request, uuid):

    participant = get_object_or_404(
        Participant,
        uuid=uuid
    )

    if request.method == "POST":
        participant.delete()
        return redirect('participant_list')


    return render(
        request,
        'participants/delete.html',
        {
            'participant': participant
        }
    )


def participant_update(request, uuid):

    participant = get_object_or_404(
        Participant,
        uuid=uuid
    )

    if request.method == "POST":

        form = ParticipantForm(
            request.POST,
            request.FILES,
            instance=participant
        )

        if form.is_valid():
            form.save()
            return redirect("participant_list")

    else:
        form = ParticipantForm(instance=participant)

    return render(
        request,
        "participants/form.html",
        {
            "form": form,
            "title": "Modifier un participant",
        },
    )
